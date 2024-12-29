from importlib import import_module
from ubexreciever import UbexReciever
from typing import List
from openpyxl import Workbook
import excelstyles01 as exs01
import datetime
from logging import Logger
import locale

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import os.path
import smtplib

class UbexOperBase:

    connname : str
    grpid : str
    qid : str
    uname: str
    e_result = []
    lst_ws = []
    reciever : UbexReciever
    folder : str
    filesuffix : str
    fname : str
    frmt : str
    logger : Logger
    isPrint : bool
    batch_ts: str

    xlsxFile: str
    xlsxFullFile: str

    wb : Workbook

    def __init__(self, connname, grpid: str, qid: str, uname: str, batch_ts: str, logger: Logger):
        self.connname = connname
        self.grpid = grpid
        self.qid = qid
        self.uname = uname
        if batch_ts:
            self.batch_ts = batch_ts
        else:
            self.batch_ts = datetime.datetime.now().strftime("%Y%m%d%H%M%S")

        self.logger = logger
        self.isPrint = True
        self.lst_ws = []
        self.reciever = UbexReciever(logger = self.logger, batch_ts = self.batch_ts)
        self.xlsxFile = ""
        self.xlsxFullFile = ""
        self.e_result = self.reciever.getGrpDetail(self.connname, self.grpid, self.qid, self.uname)
        if self.e_result:
            item = self.e_result[0]
            self.filesuffix = item['FILESUFFIX']
            self.fname      = item['FNAME']
            self.frmt       = item['FRMT']
            self.folder     = item['FOLDER']
            for item in self.e_result:
                ws = {}
                ws['QID'] = item['QID']
                ws['WSPYCLASS'] = item['WSPYCLASS']
                ws['TECHNAME'] = item['TECHNAME']
                ws['WSNAME'] = item['WSNAME']
                ws['DESCRIPTION'] = item['DESCRIPTION']
                ws['DATA'] = self.reciever.getQueryData(self.connname, self.grpid, item['QID'])
                self.lst_ws.append(ws)
            del self.e_result
        else:
            self._log_error("Ошибка запроса группы {0}".format(self.grpid))


    def _log_info(self, mess: str):
        self.logger.info(mess)
        if self.isPrint: print(mess)


    def _log_error(self, mess: str):
        self.logger.error(mess)
        if self.isPrint: print(mess) 

    @property
    def bwuser(self):
        if self.uname:
            return self.uname
        else:
            return self.reciever.bwuser


    def _getFullFileName(self) -> str:
        if not self.xlsxFile :
            if self.filesuffix == 'ts':
                # wbsuffix_value = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                wbsuffix_value = self.batch_ts
            elif self.filesuffix == 'u_ts':
                # wbsuffix_value = '{0}_{1}'.format(self.bwuser, datetime.datetime.now().strftime("%Y%m%d%H%M%S"))
                wbsuffix_value = '{0}_{1}'.format(self.bwuser, self.batch_ts)
                # wbsuffix_value = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
            elif self.filesuffix == '':
                wbsuffix_value = ''
            else:
                raise self._log_error(' Undefined wbsuffix: {}'.format(self.filesuffix))
            self.xlsxFullFile = '{0}\\{1}_{2}.xlsx'.format(self.folder, self.fname, wbsuffix_value)
            self.xlsxFile = '{0}{1}.xlsx'.format(self.fname, wbsuffix_value)
        return self.xlsxFullFile


    def _saveXSLX(self):
        # create workbook
        self.wb = Workbook()
        locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')
        self._addAnalysisStyles()

        # fill data
        ws = self.wb.active
        ws.title = self.lst_ws[0]['WSNAME']
        for wsn in self.lst_ws[1:]:
            ws = self.wb.create_sheet(title=wsn['WSNAME']) 

        # Worksheet для отображения параметров
        wsp = self.wb.create_sheet(title='Параметры')
        for item in self.lst_ws:
            ws = self.wb[item['WSNAME']]
            wspyclass = str(item['WSPYCLASS'])
            # Получение динамического имени модуля (Файла) и класса
            wspy_module, wspy_class = wspyclass.split('.')
            try:
                cla = getattr(import_module(wspy_module), wspy_class)
                wbcl = cla()
                wbcl.fill(ws, item['DATA'], wsp)
                self._log_info("Batch {1}  Лист {0} сохранен".format(item['WSNAME'], self.reciever.batch_ts))
            except (AttributeError, ModuleNotFoundError) as e:
                self._log_error("Batch {1} {0}".format(str(e), self.reciever.batch_ts))
                return      
        # fp = '{0}\\{1}{2}.xlsx'.format(self.folder, self.fname, wbsuffix_value)
        fp = self._getFullFileName()
        try:
            self.wb.save(filename = fp)
        except FileNotFoundError as e:
            self._log_error("Batch {1} {0}".format(str(e), self.reciever.batch_ts))
            return
        self._log_info("Batch {1} {0} сохранен".format(fp, self.reciever.batch_ts))

    def _getEmailSubject(self) -> str:
        if self.lst_ws:
            return self.lst_ws[0]['DESCRIPTION']
        else:
            if self.qid:
                return 'Отчет {0} в группе {1} (из {2})'.format(self.qid, self.grpid, self.connname)
            else:
                return 'Отчеты группы {0} (из {1})'.format(self.grpid, self.connname)            

    
    def _saveXLSXemail(self):
        self._saveXSLX()
        email_recipient = self.reciever.getEmails(self.connname, self.grpid, self.qid, self.uname)
        if email_recipient:
# smtp = smtplib.SMTP('mailhub.company.com') 
# smtp.ehlo()   
# max_limit_in_bytes = int( smtp.esmtp_features['size'] )
# https://stackoverflow.com/questions/64466028/check-the-full-size-of-an-email-before-sending-python-smtplib
            email_sender = os.environ.get('UBEXSENDMAIL')
            email_message = 'Ваш отчет(ы) - во вложении!'
            # if self.qid:
            #     email_subject = 'Отчет {0} в группе {1} (из {2})'.format(self.qid, self.grpid, self.connname)
            # else:
            #     email_subject = 'Отчеты группы {0} (из {1})'.format(self.grpid, self.connname)
            email_subject = self._getEmailSubject()
            
            msg = MIMEMultipart()
            msg['From'] = email_sender
            msg['To'] = ', '.join(email_recipient)
            msg['Subject'] = email_subject
            msg.attach(MIMEText(email_message, 'plain'))
            fp = self._getFullFileName()
            if fp:
                attachment = open(fp, "rb")
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition',
                            "attachment; filename= %s" % self.xlsxFile)
                msg.attach(part)
        #Сервер, порт, логин и пароль для отправки
                try :
                    server = smtplib.SMTP('smtp', 25)
                    server.ehlo()
                    server.starttls()
                    smtppassw = os.environ.get('UBEXPASSMAIL')
                    smtpuser = os.environ.get('UBEXLOGMAIL')
                    server.login(smtpuser, smtppassw)
                    text = msg.as_string()
                    server.sendmail(email_sender, email_recipient, text)
                    self._log_info('e-mail sent')
                    server.quit()
                except :
                    self._log_error("SMTP server connection error")
                return True
 

    def _addAnalysisStyles(self):
        self.wb.add_named_style(exs01.colheader_regular)
        self.wb.add_named_style(exs01.colheader_total)
        self.wb.add_named_style(exs01.rowheader_regular)
        self.wb.add_named_style(exs01.rowheader_total)
        self.wb.add_named_style(exs01.cellvalue_regular)
        self.wb.add_named_style(exs01.cellvalue_totlr)
        self.wb.add_named_style(exs01.cellvalue_totlc)


    def run(self):
        if self.frmt.lower() == 'xlsx':
            self._saveXSLX()
        elif self.frmt.lower() == 'xlsxemail':
            self._saveXLSXemail()