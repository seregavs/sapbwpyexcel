from openpyxl import worksheet
from openpyxl.styles import numbers
from typing import Dict
import excelstyles01 as exs01
from ubexwb import UbexWorkbook
import locale

class UbexWorkbook_l01( UbexWorkbook ):
   
    def fill(self, ws: worksheet, qdata: Dict, wsp: worksheet):

        # KF_first_col = 5 # Номер первой колонки с числами (считая от 0)
        celldel = ';'
        chBase = 64

        # Показываем параметры, передаваемые в отчет
        black_list  = set(("GRPID","QID", "VARTYP", "VPROCTP", "IOBJNM", "VPARSEL", "VARINPUT", "ENTRYTP"))
        qp = [ { k:v for k, v in li.items() if k not in black_list } for li in qdata['E_QPARAM']]

        wsparam_list = [d['LOW'] for d in qp if d['PARAMNAME'] == '_WSPARAM']
        if wsparam_list:
            topL, KF_first_col, topLfix, *rest = wsparam_list[0].split(',')
        else:
            topL, KF_first_col, topLfix = 'A5', '5', ''
        if KF_first_col:
            KF_first_col = int(KF_first_col)
        else:
            KF_first_col = 5
        if not topL:
            topL = 'A5'
        topLfix = topLfix.strip()    
        begC, begR = int(ord(topL[0])-chBase), int(topL[1:]) # Column и Rows начинаются с 1
       
        # ws['A1'] = str(qp)
        self.fill_param(wsp, qp, ws.title)

        e_status = str(qdata['E_STATUS'])
        if  e_status != '0':
            ws['A2'] = e_status
            return
        dims = {} # for storing columns widths
        data = []
        for r in qdata['E_RESULT']:
            data.append(r['REPROW'][:-1])

        if not data:
            return
        
        header = [x.strip() for x in data[0].split(celldel)]
        for x, hi in enumerate(header):
            (d := ws.cell(row=begR, column=begC+x, value=hi)).style = exs01.colheader_total
        for y, r in enumerate(data[1:]):
            rlst = [x.strip() for x in r.split(celldel)]
            for x, c in enumerate(rlst):
                if x >= KF_first_col:
                   d = ws.cell(row=y+begR+1, column=begC+x, value=locale.atof(c))
                else:
                   c = c[1:] if c.startswith("'") else c
                   d = ws.cell(row=y+begR+1, column=begC+x, value=c)
                d.style = exs01.rowheader_regular if x < KF_first_col else exs01.cellvalue_regular
                if x >= KF_first_col:
                    d.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
                    dims[d.column_letter] = max((dims.get(d.column_letter, 0), len(str(d.value))+4)) 
                else:
                    dims[d.column_letter] = max((dims.get(d.column_letter, 0), len(str(d.value))+0))

        # (d := ws.cell(row=3+len(data), column=8+2, value="=SUM(J{0}:J{1})".format(4,3+len(data)-1))).style = exs01.cellvalue_totlc
        # d.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED2
        # dims[d.column_letter] = max((dims.get(d.column_letter, 0), len(str(d.value)))) 

        for col, value in dims.items():
            ws.column_dimensions[col].width = value+2

        filters = ws.auto_filter
        filters.ref = "{0}{1}:{2}{3}".format(chr(chBase+begC), begR, chr(chBase+begC+2), begR+1)

        ws.freeze_panes = ws['{0}{1}'.format(chr(chBase+begC+1), begR+1)]

        # ws.sheet_properties.outlinePr.summaryRight = True
        # ws.column_dimensions.group('E','E',outline_level=1)
        # ws.column_dimensions.group('I','I',outline_level=1)
        if topLfix:
            ws.freeze_panes = ws['{0}'.format(topLfix)]
            # ws.freeze_panes = ws['{0}{1}'.format(chr(chBase+begC+1), begR+1)]