from openpyxl import worksheet
from openpyxl.styles import numbers
from typing import List
import excelstyles01 as exs01

class UbexWorkbook:

    def fill_param(self, wsp: worksheet, qp: List, wstitle: str):
            param_header = ['Лист', '№наЛисте', 'Параметр', \
                            'Параметр(имя)', 'Счетчик', 'Вкл', \
                            'Тип', 'Нижнее знач','Верхнее знач']
            dims = {} # for storing columns widths
            dims['A'] = 10
            dims['B'] = 5
            dims['C'] = 18
            dims['D'] = 50
            dims['E'] = 5        
            dims['F'] = 5
            dims['G'] = 5
            dims['H'] = 30
            dims['I'] = 10
            rs = 2
            if  wsp['A1'].value is None:
                for x, hi in enumerate(param_header):
                    (d := wsp.cell(row=1, column=1+x, value=hi)).style = exs01.colheader_total
                filters = wsp.auto_filter
                filters.ref = "A1:I1"
                for col, value in dims.items():
                    wsp.column_dimensions[col].width = value
                wsp.freeze_panes = wsp['D2']
            else:
                while not (wsp['A{0}'.format(rs)].value is None):
                    rs += 1

            for item in qp:
                d = wsp.cell(row=rs, column=1, value=wstitle)
                d = wsp.cell(row=rs, column=2, value='1')
                d = wsp.cell(row=rs, column=3, value=item['PARAMNAME'])
                d = wsp.cell(row=rs, column=4, value=item['PARAMDESCR'])
                d = wsp.cell(row=rs, column=5, value=item['CNT'])
                d = wsp.cell(row=rs, column=6, value=item['SIGN'])
                d = wsp.cell(row=rs, column=7, value=item['OPTI'])
                d = wsp.cell(row=rs, column=8, value=item['LOW'])
                d = wsp.cell(row=rs, column=9, value=item['HIGH'])
                rs += 1