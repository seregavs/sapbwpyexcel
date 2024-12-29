from openpyxl.styles import NamedStyle, Font, Border, Side, Color, PatternFill, Alignment

bd = Side(style='thin', color="000000") # тонкий
colheader_regular = NamedStyle(name="colheader_regular",
                # font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'center', vertical = 'center'),                
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF9AD4FF'))
                )
colheader_total = NamedStyle(name="colheader_total",
                font = Font(name = "Verdana", size="8" ),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'center', vertical = 'center'),
                # fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF7BAACC'))
                fill = PatternFill(patternType='solid', fgColor=Color('FF%02x%02x%02x' % (183,207,232))) 
                )
rowheader_total = NamedStyle(name="rowheader_total",
                font = Font(bold=True, name = "Verdana", size="8" ),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'left', vertical = 'center'),
                fill = PatternFill(patternType='solid', fgColor=Color('FF%02x%02x%02x' % (219,229,241))) 
                # fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF9AD4FF')) '#%02x%02x%02x' % (219,229,241)
                )    
rowheader_regular = NamedStyle(name="rowheader_regular",
                font = Font(name = "Verdana", size="8" ),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'left', vertical = 'center'),
                # fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF9AD4FF'))
                fill = PatternFill(patternType='solid', fgColor=Color('FF%02x%02x%02x' % (219,229,241))) 
                )              
rowheader_total2 = NamedStyle(name="rowheader_total2",
                # font = Font(bold=True),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'left'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FF7BAACC'))
                )  
cellvalue_totlr = NamedStyle(name="cellvalue_totlr",
                font = Font(bold=True, size = "8", name = "Verdana"),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'right'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFE8E24A'))
                )
cellvalue_totlc = NamedStyle(name="cellvalue_totlc",
                font = Font(bold=True, size="8", name = "Verdana"),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'right'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFEAED96'))
                )                
cellvalue_regular = NamedStyle(name="cellvalue_regular",
                font = Font(size="8", name = "Verdana"),
                border = Border(left=bd, top=bd, right=bd, bottom=bd),
                alignment = Alignment(horizontal = 'right'),
                fill = PatternFill(patternType='solid', fgColor=Color(rgb='FFFFFFFF'))
                )